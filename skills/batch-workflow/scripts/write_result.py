#!/usr/bin/env python3
"""Write one task result back to an output file (Excel or JSON).

Designed to be called once per task during batch-workflow processing.
Saves after every call so an interrupted run leaves a fully-consistent file
up to the last completed task.

Behaviour:
- status=found: writes data fields into the output row; sets _status column to "found"
- status=no_result: writes NOTHING to data columns; sets _status to "no_result"
- status=error: writes NOTHING to data columns; sets _status to "error";
  writes --error-msg to _error column if present
"""

import argparse
import json
import sys
from pathlib import Path


# Column headers managed by this script (always added if missing)
_STATUS_COL = "_status"
_ERROR_COL = "_error"


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------


def _get_or_create_workbook(output_path: Path):
    try:
        from openpyxl import load_workbook, Workbook
    except ImportError:
        sys.exit("openpyxl is required: uv add openpyxl")

    if output_path.exists():
        return load_workbook(output_path)
    wb_new = Workbook()
    wb_new.active.title = "Results"
    return wb_new


def _ensure_header(ws, col_name: str) -> int:
    """Return 1-based column index for col_name, creating it if missing."""
    for col_idx in range(1, (ws.max_column or 0) + 2):
        cell_val = ws.cell(row=1, column=col_idx).value
        if cell_val == col_name:
            return col_idx
        if cell_val is None:
            ws.cell(row=1, column=col_idx, value=col_name)
            return col_idx
    # Shouldn't reach here, but append defensively
    new_col = (ws.max_column or 0) + 1
    ws.cell(row=1, column=new_col, value=col_name)
    return new_col


def write_excel(output_path: Path, task_id: int, data: dict | None, status: str, error_msg: str | None) -> None:
    wb = _get_or_create_workbook(output_path)
    ws = wb.active
    row = task_id + 1  # row 1 = header, task_id is 1-based data row

    if data and status == "found":
        for col_name, value in data.items():
            col_idx = _ensure_header(ws, col_name)
            ws.cell(row=row, column=col_idx, value=value)

    status_col = _ensure_header(ws, _STATUS_COL)
    ws.cell(row=row, column=status_col, value=status)

    if error_msg:
        error_col = _ensure_header(ws, _ERROR_COL)
        ws.cell(row=row, column=error_col, value=error_msg)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# JSON helpers
# ---------------------------------------------------------------------------


def write_json(output_path: Path, task_id: int, data: dict | None, status: str, error_msg: str | None) -> None:
    results: list[dict] = []
    if output_path.exists():
        try:
            results = json.loads(output_path.read_text(encoding="utf-8"))
            if not isinstance(results, list):
                results = []
        except Exception:
            results = []

    # Find existing entry for this task_id or append
    entry_idx = next((i for i, r in enumerate(results) if r.get("_task_id") == task_id), None)
    entry: dict = {"_task_id": task_id, "_status": status}
    if data and status == "found":
        entry.update(data)
    if error_msg:
        entry["_error"] = error_msg

    if entry_idx is not None:
        results[entry_idx] = entry
    else:
        results.append(entry)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(results, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description="Write one task result to output file (Excel or JSON)")
    parser.add_argument("--output", required=True, help="Output file path (.xlsx or .json)")
    parser.add_argument("--task-id", required=True, type=int, help="1-based task/row number")
    parser.add_argument("--data", default=None, help="Result data as a JSON object string")
    parser.add_argument(
        "--status",
        choices=["found", "no_result", "error"],
        default="found",
        help="Task outcome (default: found)",
    )
    parser.add_argument("--error-msg", default=None, help="Error description (used when --status=error)")
    args = parser.parse_args()

    out_path = Path(args.output)
    suffix = out_path.suffix.lower()

    data: dict | None = None
    if args.data:
        try:
            data = json.loads(args.data)
            if not isinstance(data, dict):
                sys.exit(f"--data must be a JSON object, got: {type(data)}")
        except json.JSONDecodeError as exc:
            sys.exit(f"Invalid JSON in --data: {exc}")

    if args.status == "no_result":
        data = None  # never write data columns for no_result

    if suffix in {".xlsx", ".xlsm"}:
        write_excel(out_path, args.task_id, data, args.status, args.error_msg)
    else:
        write_json(out_path, args.task_id, data, args.status, args.error_msg)

    print(json.dumps({"task_id": args.task_id, "status": args.status, "output": str(out_path)}))


if __name__ == "__main__":
    main()
