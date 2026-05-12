#!/usr/bin/env python3
"""Load and inspect a task list from Excel, CSV, JSON, or plain-text files.

Outputs a JSON summary to stdout so the agent can understand the task structure,
identify resume position, and plan the workflow template.
"""

import argparse
import csv
import json
import sys
from pathlib import Path


def _load_checkpoint_index(checkpoint_path: Path) -> int:
    """Return the 1-based index of the first incomplete task, or 1 if no checkpoint."""
    if not checkpoint_path.exists():
        return 1
    try:
        data = json.loads(checkpoint_path.read_text(encoding="utf-8"))
        completed = set(data.get("completed") or [])
        total = data.get("total") or 0
        if not completed or not total:
            return 1
        # Find first task_id (1-based) not in completed
        for i in range(1, total + 1):
            if i not in completed and str(i) not in completed:
                return i
        return total + 1  # all done
    except Exception:
        return 1


def _load_excel(path: Path) -> tuple[list[str], list[dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl is required for Excel files: uv add openpyxl")

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], []

    # First row is header
    headers = [str(cell).strip() if cell is not None else f"column_{i+1}" for i, cell in enumerate(rows[0])]
    tasks = []
    for row_idx, row in enumerate(rows[1:], start=2):
        task: dict = {"_row": row_idx}
        for col_idx, (header, value) in enumerate(zip(headers, row)):
            task[header] = value
        tasks.append(task)
    return headers, tasks


def _load_csv(path: Path) -> tuple[list[str], list[dict]]:
    tasks = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames or [])
        for row_idx, row in enumerate(reader, start=2):
            task = {"_row": row_idx}
            task.update(dict(row))
            tasks.append(task)
    return headers, tasks


def _load_json(path: Path) -> tuple[list[str], list[dict]]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(raw, list):
        items = raw
    elif isinstance(raw, dict) and "tasks" in raw:
        items = raw["tasks"]
    else:
        items = [raw]

    tasks = []
    headers: list[str] = []
    for idx, item in enumerate(items, start=1):
        if isinstance(item, dict):
            task = {"_row": idx}
            task.update(item)
            for k in item:
                if k not in headers:
                    headers.append(k)
        else:
            task = {"_row": idx, "content": str(item)}
            if "content" not in headers:
                headers.append("content")
        tasks.append(task)
    return headers, tasks


def _load_text(path: Path) -> tuple[list[str], list[dict]]:
    lines = [line.rstrip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
    tasks = [{"_row": idx, "content": line} for idx, line in enumerate(lines, start=1)]
    return ["content"], tasks


def load_tasks(input_path: Path, checkpoint_path: Path | None = None) -> dict:
    suffix = input_path.suffix.lower()

    if suffix in {".xlsx", ".xlsm", ".xls"}:
        fields, tasks = _load_excel(input_path)
    elif suffix == ".csv":
        fields, tasks = _load_csv(input_path)
    elif suffix == ".json":
        fields, tasks = _load_json(input_path)
    elif suffix in {".txt", ".md", ""}:
        fields, tasks = _load_text(input_path)
    else:
        # Try text as fallback
        fields, tasks = _load_text(input_path)

    total = len(tasks)
    resume_index = 1
    if checkpoint_path:
        resume_index = _load_checkpoint_index(checkpoint_path)

    samples = []
    for task in tasks[:3]:
        # Exclude internal _row from samples to keep output readable
        samples.append({k: v for k, v in task.items() if k != "_row"})

    return {
        "file": str(input_path),
        "format": suffix.lstrip(".") or "text",
        "total": total,
        "fields": fields,
        "samples": samples,
        "resume_index": resume_index,
        "all_done": resume_index > total,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Load and inspect a task list for batch-workflow processing")
    parser.add_argument("--input", required=True, help="Path to task list file (Excel, CSV, JSON, or text)")
    parser.add_argument("--checkpoint", help="Optional checkpoint JSON path for resume detection")
    parser.add_argument("--output", help="Optional path to write JSON report instead of stdout")
    args = parser.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        sys.exit(f"Input file not found: {in_path}")

    cp_path = Path(args.checkpoint) if args.checkpoint else None
    report = load_tasks(in_path, cp_path)
    rendered = json.dumps(report, ensure_ascii=False, indent=2, default=str)

    if args.output:
        out_path = Path(args.output)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(rendered, encoding="utf-8")
        print(f"Saved task report to {out_path}")
    else:
        print(rendered)


if __name__ == "__main__":
    main()
