#!/usr/bin/env python3
"""Inspect workbook structure for Excel modeling workflows."""

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


TYPE_ORDER = ["number", "text", "boolean", "date", "empty", "other"]


def _classify_value(value: object) -> str:
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if hasattr(value, "isoformat"):
        return "date"
    if isinstance(value, str):
        return "text"
    return "other"


def inspect_workbook(path: Path) -> dict:
    wb = load_workbook(path, data_only=False, read_only=True)
    output: dict[str, object] = {
        "file": str(path),
        "sheet_count": len(wb.sheetnames),
        "sheets": [],
    }

    for name in wb.sheetnames:
        ws = wb[name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        formula_count = 0
        header: list[str] = []
        column_profiles: dict[str, dict[str, int]] = {}

        for c in range(1, max_col + 1):
            header_value = ws.cell(row=1, column=c).value
            label = str(header_value).strip() if header_value is not None else f"column_{c}"
            if not label:
                label = f"column_{c}"
            header.append(label)
            column_profiles[label] = {k: 0 for k in TYPE_ORDER}

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                if c <= len(header):
                    key = header[c - 1]
                    kind = _classify_value(value)
                    column_profiles[key][kind] = column_profiles[key].get(kind, 0) + 1

        output["sheets"].append(
            {
                "name": name,
                "max_row": max_row,
                "max_column": max_col,
                "formulas": formula_count,
                "header": header,
                "column_profiles": column_profiles,
            }
        )

    wb.close()
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect workbook structure for advanced modeling")
    parser.add_argument("--input", required=True, help="Path to input workbook (.xlsx/.xlsm)")
    parser.add_argument("--output", required=False, help="Optional output JSON path")
    args = parser.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        raise SystemExit(f"Input workbook not found: {in_path}")

    report = inspect_workbook(in_path)
    rendered = json.dumps(report, ensure_ascii=False, indent=2)

    if args.output:
        out_path = Path(args.output)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(rendered, encoding="utf-8")
        print(f"Saved inspection report to {out_path}")
    else:
        print(rendered)


if __name__ == "__main__":
    main()
