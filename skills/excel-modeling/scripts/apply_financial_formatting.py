#!/usr/bin/env python3
"""Apply basic blue/black/green financial modeling font conventions."""

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

BLUE = "000000FF"
BLACK = "00000000"
GREEN = "00008000"


def apply_styles(input_path: Path, output_path: Path) -> tuple[int, int, int]:
    wb = load_workbook(input_path)
    blue_count = 0
    black_count = 0
    green_count = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value

                # Formula cells: black
                if isinstance(value, str) and value.startswith("="):
                    cell.font = Font(color=BLACK)
                    black_count += 1
                    continue

                # Hyperlink/reference cells: green
                if cell.hyperlink is not None:
                    cell.font = Font(color=GREEN, underline="single")
                    green_count += 1
                    continue

                # Hardcoded non-empty scalar values: blue
                if value is not None and isinstance(value, (int, float, str, bool)):
                    cell.font = Font(color=BLUE)
                    blue_count += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return blue_count, black_count, green_count


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply financial-model formatting conventions")
    parser.add_argument("--input", required=True, help="Path to input workbook")
    parser.add_argument("--output", required=True, help="Path to output workbook")
    args = parser.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)

    if not in_path.exists():
        raise SystemExit(f"Input workbook not found: {in_path}")

    blue_count, black_count, green_count = apply_styles(in_path, out_path)
    print(
        "Applied formatting: "
        f"blue_inputs={blue_count}, black_formulas={black_count}, green_links={green_count}. "
        f"Saved to {out_path}"
    )


if __name__ == "__main__":
    main()
