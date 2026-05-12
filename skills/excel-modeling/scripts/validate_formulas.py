#!/usr/bin/env python3
"""Validate workbook formulas and detect common spreadsheet error cells."""

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

ERROR_LITERALS = {"#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"}


def _address(row: int, col: int) -> str:
    from openpyxl.utils import get_column_letter

    return f"{get_column_letter(col)}{row}"


def validate(path: Path) -> dict:
    raw_wb = load_workbook(path, data_only=False, read_only=True)
    value_wb = load_workbook(path, data_only=True, read_only=True)

    report: dict[str, object] = {
        "file": str(path),
        "sheets": [],
        "summary": {
            "formula_cells": 0,
            "error_cells": 0,
        },
    }

    for name in raw_wb.sheetnames:
        raw_ws = raw_wb[name]
        value_ws = value_wb[name]
        sheet_formulas = []
        sheet_errors = []

        max_row = raw_ws.max_row or 0
        max_col = raw_ws.max_column or 0
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                raw_cell = raw_ws.cell(row=r, column=c)
                value_cell = value_ws.cell(row=r, column=c)
                raw_value = raw_cell.value

                if isinstance(raw_value, str) and raw_value.startswith("="):
                    sheet_formulas.append(
                        {
                            "cell": _address(r, c),
                            "formula": raw_value,
                        }
                    )

                computed = value_cell.value
                if isinstance(computed, str) and computed in ERROR_LITERALS:
                    sheet_errors.append(
                        {
                            "cell": _address(r, c),
                            "error": computed,
                            "formula": raw_value if isinstance(raw_value, str) and raw_value.startswith("=") else None,
                        }
                    )

        report["summary"]["formula_cells"] += len(sheet_formulas)
        report["summary"]["error_cells"] += len(sheet_errors)
        report["sheets"].append(
            {
                "name": name,
                "formula_count": len(sheet_formulas),
                "error_count": len(sheet_errors),
                "formulas": sheet_formulas,
                "errors": sheet_errors,
            }
        )

    raw_wb.close()
    value_wb.close()
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate formulas and error cells in workbook")
    parser.add_argument("--input", required=True, help="Path to input workbook")
    parser.add_argument("--output", required=False, help="Optional output JSON path")
    args = parser.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        raise SystemExit(f"Input workbook not found: {in_path}")

    report = validate(in_path)
    rendered = json.dumps(report, ensure_ascii=False, indent=2)

    if args.output:
        out_path = Path(args.output)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(rendered, encoding="utf-8")
        print(f"Saved validation report to {out_path}")
    else:
        print(rendered)


if __name__ == "__main__":
    main()
